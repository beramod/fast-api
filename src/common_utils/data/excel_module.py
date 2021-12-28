import openpyxl
from config import BASE_DIR


class ExcelModule:
    def __init__(self, file_name):
        self._file_name = file_name
        self._wb = openpyxl.Workbook()
        self._ws = self._wb.active
        self._offset = 1

    def get_cur_work_sheet(self):
        return self._ws

    def set_line_alignment(self):
        self._offset += 1

    def set_titles(self, titles: list):
        for idx, title in enumerate(titles):
            self._ws.cell(row=self._offset, column=idx + 1, value=title)

    def create_sheet(self, sheet_name):
        self._ws = self._wb.create_sheet(sheet_name)

    def remove_sheet(self, sheet_name):
        self._wb.remove_sheet(self._wb[sheet_name])

    def select_sheet(self, sheet_name):
        self._ws = self._wb[sheet_name]

    def set_contents(self, contents: list):
        for idx1, content in enumerate(contents):
            for idx2, each in enumerate(content):
                self._ws.cell(row=idx1 + 2, column=idx2 + 1, value=each) \

    def save(self):
        path = '{}/files/{}.xlsx'.format(BASE_DIR, self._file_name)
        self._wb.save(path)
        return path